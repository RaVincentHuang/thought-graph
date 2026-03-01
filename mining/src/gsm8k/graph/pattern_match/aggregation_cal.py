import networkx as nx
import json
import re
import os
import argparse
from collections import deque
from openpyxl import Workbook

def calculate_to_tag(G, root, tag):
    def get_path_length(node): 
        if G.out_degree(node) == 0:
            return 0
        return 1 + max((get_path_length(succ) for succ in G.successors(node)), default=0)
   
    def get_distance_to_root(node):
        dist = 0
        curr = node
        while curr != root:
            dist += 1
            curr = next(G.predecessors(curr))
        return dist
   
    def dfs(node, current_prob):
        if G.nodes[node]['tag'] == tag:
            match_count = G.nodes[node].get('match_count', 0)
            
            total_path_length = get_distance_to_root(node) + get_path_length(node)
            weight = match_count / total_path_length if total_path_length > 0 else match_count
            return current_prob * weight
           
        total_prob = 0
        for neighbor in G.successors(node):
            edge_prob = G[node][neighbor].get('probability', 0)
            total_prob += dfs(neighbor, current_prob * edge_prob)
        return total_prob
    return dfs(root, 1.0)
    
def gen_prob_graph(G):
    prob_graph = G.copy()
    
    for node in prob_graph.nodes():
        prob_graph.nodes[node]['match_count'] = 0

    
    leaves = [n for n in prob_graph.nodes() if prob_graph.out_degree(n) == 0]
    
    for leaf in leaves:
        node = leaf
        while node != 0:
            prob_graph.nodes[node]['success'] = 1
            node = next(prob_graph.predecessors(node))
    
    queue = deque([0])
    while queue:
        node = queue.popleft()
        for neighbor in prob_graph.successors(node):
            
            prob_graph.nodes[neighbor]['match_count'] = prob_graph.nodes[node]['match_count']
            high_avg_score = False
            acc = float(prob_graph.nodes[neighbor].get('acc', 0))
            logi = float(prob_graph.nodes[neighbor].get('logi', 0))
            
            infer = float(prob_graph.nodes[neighbor].get('infer', 0))
            if acc >= 7.77 and logi >= 7.36 and infer >= 6.63:
                high_avg_score = True

            if high_avg_score:
                prob_graph.nodes[neighbor]['match_count'] += 1
            
            
            queue.append(neighbor)
    for node in prob_graph.nodes():
        neighbors = list(prob_graph.successors(node))
        if neighbors:
            weights = [prob_graph[node][neighbor]['value'] for neighbor in neighbors]
            total_weight = sum(weights)
            for neighbor in neighbors:
                
                prob_graph[node][neighbor]['probability'] = prob_graph[node][neighbor]['value'] / total_weight

    return prob_graph

def get_answer_path(log_path):
    with open(log_path,"r",encoding="utf-8") as file:
        log_data = file.read()
    output_pattern = re.compile(r"output='([^']*)'")
    plan_pattern = re.compile(r"'plan': '([^']*)'")
    correct_pattern = re.compile(r"correct=(\w+)")

    outputs = output_pattern.findall(log_data)
    plans = plan_pattern.findall(log_data)
    corrects = correct_pattern.findall(log_data)
    
    new_plans = []
    for plan in plans:
        tmp = plan[2:].split("\\n")[:-2]
        new_plans.append(tmp)
    return outputs, new_plans,corrects

def assign_levels(G, root):
    levels = {root: 0}
    queue = deque([root])
    while queue:
        node = queue.popleft()
        current_level = levels[node]
        for neighbor in G.neighbors(node):
            if neighbor not in levels:  
                levels[neighbor] = current_level + 1
                queue.append(neighbor)
    return levels

def load_config(config_path):
    with open(config_path, 'r') as f:
        return json.load(f)


def model_mapping(model):
    if model == "gpt35":
        return "gpt-3.5-turbo"
    elif model == "qwen15_7b":
        return "qwen-1.5-7b"
    elif model == "mistral_7b":
        return "mistral-7b"
    elif model == "glm4_9b":
        return "glm-4-9b"
    elif model == "llama3_8b":
        return "llama-3-8b"

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process data for a specific experiment.")
    parser.add_argument('--exp_name', type=str, required=True, help='The name of the experiment.')
    parser.add_argument('--benchmark_model', type=str, required=True, help='The benchmark model to use.')
    parser.add_argument('--cluster_id', type=int, required=True, help='The cluster ID to process.')
    parser.add_argument('--config', type=str, required=True, help='Path to the configuration file.')
    parser.add_argument('--mode', type=str, required=True, help='The mode to use.')
    args = parser.parse_args()
    
    model_list = load_config(args.config).get("model_list", [])
    benchmark_model = args.benchmark_model
    exp_type = args.exp_name
    cluster_num = args.cluster_id
    mode = args.mode
    
    if exp_type == "exp1" or exp_type == "exp2":
        MODEL_ORDER = [
            "gpt35",
            "glm4_9b",
            "llama3_8b",
            "qwen15_7b",
            "mistral_7b"
        ]
    
    model_data = {}
    
    for model in model_list:
        input_path = rf"processed_data/gsm8k/{exp_type}_results_{benchmark_model}/{mode}_graph_{cluster_num}/{model}"
        result_log_path = rf"raw_data/gsm8k/{exp_type}/result_log/{model}/result.log"
        save_path = rf"processed_data/gsm8k/{exp_type}_results/{model}_graph_by_instance"
        
        
        outputs, plans,corrects = get_answer_path(result_log_path)
        tag_aggregation_dict = {}  
        count_tmp = 0
        for file in range(len(corrects)):
            
            file = input_path + f"/tree_{file}.txt"
            G = nx.DiGraph()
            

            with open(file, "r", encoding="utf-8",errors='ignore') as f:
                for line in f:
                    line = line.encode('utf-8', 'ignore').decode('utf-8')
                    parts = line.split()
                    if parts[0] == "v":
                        G.add_node(int(parts[1]), tag=int(parts[2]), success=int(parts[3]), acc=float(parts[4]), logi=float(parts[5]), dis=float(parts[6]), infer=float(parts[7]))
                    elif parts[0] == "e":
                        G.add_edge(int(parts[1]), int(parts[2]), value=float(parts[3]),action=str(" ".join(parts[4:])))

            root_node = 0
            
            G_prob = gen_prob_graph(G)
            leaves = [n for n in G.nodes() if G.out_degree(n) == 0]
            
            max_match_count = max(G_prob.nodes[leaf]['match_count'] for leaf in leaves)
            if max_match_count:
                leaves = [leaf for leaf in leaves if G_prob.nodes[leaf]['match_count'] == max_match_count]
            else:
                
                leaves = []
            count_tmp += len(leaves)

            
            leaves_tag = list(set([G_prob.nodes[leaf]['tag'] for leaf in leaves if G_prob.nodes[leaf]['tag'] not in [-1, -2]]))
            
            if len(leaves_tag) > 0:
                tag_probs = {tag: calculate_to_tag(G_prob, root_node, tag) for tag in leaves_tag}
                total_sum = sum(tag_probs.values())
                
                
                if total_sum > 0:
                    for tag, prob in tag_probs.items():
                        normalized_prob = prob / total_sum
                        if tag in tag_aggregation_dict:
                            tag_aggregation_dict[tag].append(normalized_prob)
                        else:
                            tag_aggregation_dict[tag] = [normalized_prob]
                
        tag_avg_aggregation_dict = {tag: round(sum(aggregations) / 200, 2) for tag, aggregations in sorted(tag_aggregation_dict.items())}
        
        model_data[model] = {
            'tag_dict': tag_avg_aggregation_dict,
            'max_value': max(tag_avg_aggregation_dict.values()) if len(tag_avg_aggregation_dict) > 0 else 0
        }
        print(model, tag_avg_aggregation_dict)
        print(max(tag_avg_aggregation_dict.values()))
    
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Analysis"
    
    
    ws['A1'] = "Model"
    ws['B1'] = "Tag Probabilities"
    ws['C1'] = "Max Probability"
    
    
    row = 2
    for model_name in MODEL_ORDER:
        if model_name in model_data:
            try:
                ws[f'A{row}'] = model_mapping(model_name)
                ws[f'B{row}'] = str(model_data[model_name]['tag_dict'])
                ws[f'C{row}'] = model_data[model_name]['max_value']
                row += 1
                print(f"Successfully wrote data for model: {model_name}")
            except Exception as e:
                print(f"Error writing data for model {model_name}: {str(e)}")
    
    save_path = f'processed_data/gsm8k/{exp_type}_results_{benchmark_model}/aggregation_{cluster_num}.xlsx'
    try:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        wb.save(save_path)
        print(f"Successfully saved Excel file to: {save_path}")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
